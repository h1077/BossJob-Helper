"""
岗位追踪 API — 收集、列表、统计、导出
"""
import csv
import io
import json
from datetime import datetime
from flask import Blueprint, request, jsonify, Response
from models import db, InterestedJob
from utils.auth import login_required

jobs_bp = Blueprint('jobs', __name__, url_prefix='/api/jobs')


@jobs_bp.route('/collect', methods=['POST'])
@login_required
def collect_job():
    """收集岗位（脚本端调用）— upsert 模式"""
    user_id = request.current_user.id
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400

    job_id = data.get('jobId') or data.get('id') or ''
    if not job_id:
        return jsonify({'success': False, 'error': '缺少岗位ID'}), 400

    existing = InterestedJob.query.filter_by(job_id=str(job_id), user_id=user_id).first()
    if existing:
        for field, col in [
            ('companyName', 'company_name'), ('jobName', 'job_name'),
            ('salary', 'salary'), ('location', 'location'),
            ('experience', 'experience'), ('education', 'education'),
            ('jd', 'jd'), ('companyStage', 'company_stage'),
            ('companyScale', 'company_scale'), ('companyIndustry', 'company_industry'),
            ('businessInfo', 'business_info'), ('jobResponsibilities', 'job_responsibilities'),
            ('jobRequirements', 'job_requirements'), ('sourceUrl', 'source_url'),
            ('hrName', 'hr_name'), ('hrTitle', 'hr_title'),
            ('customGreeting', 'custom_greeting'),
        ]:
            if field in data:
                setattr(existing, col, data[field])
        existing.status = data.get('status', existing.status)
        existing.collected_at = data.get('collectedAt', int(datetime.utcnow().timestamp() * 1000))
        db.session.commit()
        return jsonify({'success': True, 'data': {'id': existing.id, 'updated': True}})

    job = InterestedJob(
        user_id=user_id,
        job_id=str(job_id),
        company_name=data.get('companyName', ''),
        job_name=data.get('jobName', ''),
        salary=data.get('salary', ''),
        location=data.get('location', ''),
        experience=data.get('experience', ''),
        education=data.get('education', ''),
        jd=data.get('jd', ''),
        job_responsibilities=data.get('jobResponsibilities', ''),
        job_requirements=data.get('jobRequirements', ''),
        company_stage=data.get('companyStage', ''),
        company_scale=data.get('companyScale', ''),
        company_industry=data.get('companyIndustry', ''),
        business_info=data.get('businessInfo', {}),
        source_url=data.get('sourceUrl', ''),
        status=data.get('status', 'applied'),
        hr_name=data.get('hrName', ''),
        hr_title=data.get('hrTitle', ''),
        collected_at=data.get('collectedAt', int(datetime.utcnow().timestamp() * 1000)),
        match_score=0,
        match_level='计算中',
    )
    db.session.add(job)
    db.session.commit()
    return jsonify({'success': True, 'data': {'id': job.id, 'jobId': job.job_id}})


@jobs_bp.route('/interested', methods=['GET'])
@login_required
def get_interested_jobs():
    user_id = request.current_user.id
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    status = request.args.get('status', '')
    keyword = request.args.get('keyword', '')

    query = InterestedJob.query.filter_by(user_id=user_id)
    if status:
        query = query.filter_by(status=status)
    if keyword:
        query = query.filter(
            db.or_(
                InterestedJob.company_name.contains(keyword),
                InterestedJob.job_name.contains(keyword),
            )
        )

    total = query.count()
    jobs = query.order_by(InterestedJob.collected_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify({
        'success': True,
        'data': {
            'jobs': [j.to_dict() for j in jobs.items],
            'total': total,
            'page': page,
            'per_page': per_page,
            'pages': jobs.pages,
        }
    })


@jobs_bp.route('/stats', methods=['GET'])
@login_required
def get_stats():
    user_id = request.current_user.id
    total = InterestedJob.query.filter_by(user_id=user_id).count()
    applied = InterestedJob.query.filter_by(user_id=user_id, status='applied').count()
    interviewing = InterestedJob.query.filter_by(user_id=user_id, status='interviewing').count()
    rejected = InterestedJob.query.filter_by(user_id=user_id, status='rejected').count()
    interested = InterestedJob.query.filter_by(user_id=user_id, status='interested').count()

    return jsonify({
        'success': True,
        'data': {
            'total': total,
            'applied': applied,
            'interviewing': interviewing,
            'rejected': rejected,
            'interested': interested,
        }
    })


@jobs_bp.route('/daily-stats', methods=['GET'])
@login_required
def daily_stats():
    """最近 N 天每日投递数量"""
    from sqlalchemy import func
    user_id = request.current_user.id
    days = request.args.get('days', 14, type=int)

    cutoff = datetime.utcnow().timestamp() * 1000 - days * 86400000

    results = db.session.query(
        func.date(InterestedJob.created_at),
        func.count(InterestedJob.id),
        InterestedJob.status,
    ).filter(
        InterestedJob.user_id == user_id,
        InterestedJob.collected_at >= cutoff,
    ).group_by(
        func.date(InterestedJob.created_at),
        InterestedJob.status,
    ).order_by(func.date(InterestedJob.created_at)).all()

    daily = {}
    for date_val, count, status in results:
        d = str(date_val)
        if d not in daily:
            daily[d] = {'applied': 0, 'interviewing': 0, 'rejected': 0}
        daily[d][status] = count

    return jsonify({'success': True, 'data': daily})


@jobs_bp.route('/<int:rec_id>', methods=['PUT'])
@login_required
def update_job(rec_id):
    user_id = request.current_user.id
    job = InterestedJob.query.filter_by(id=rec_id, user_id=user_id).first()
    if not job:
        return jsonify({'success': False, 'error': '岗位不存在'}), 404

    data = request.get_json(silent=True) or {}
    for field, col in [
        ('companyName', 'company_name'), ('jobName', 'job_name'),
        ('salary', 'salary'), ('location', 'location'),
        ('status', 'status'), ('matchScore', 'match_score'),
        ('matchLevel', 'match_level'), ('matchReasons', 'match_reasons'),
        ('matchedSkills', 'matched_skills'), ('matchDetails', 'match_details'),
    ]:
        if field in data:
            setattr(job, col, data[field])
    db.session.commit()
    return jsonify({'success': True, 'data': job.to_dict()})


@jobs_bp.route('/<int:rec_id>/status', methods=['PUT'])
@login_required
def update_job_status(rec_id):
    user_id = request.current_user.id
    job = InterestedJob.query.filter_by(id=rec_id, user_id=user_id).first()
    if not job:
        return jsonify({'success': False, 'error': '岗位不存在'}), 404

    data = request.get_json(silent=True) or {}
    new_status = data.get('status')
    if new_status not in ('applied', 'interviewing', 'rejected', 'interested'):
        return jsonify({'success': False, 'error': '无效的状态值'}), 400
    job.status = new_status
    db.session.commit()
    return jsonify({'success': True, 'data': {'id': job.id, 'status': job.status}})


@jobs_bp.route('/<int:rec_id>', methods=['DELETE'])
@login_required
def delete_job(rec_id):
    user_id = request.current_user.id
    job = InterestedJob.query.filter_by(id=rec_id, user_id=user_id).first()
    if not job:
        return jsonify({'success': False, 'error': '岗位不存在'}), 404
    db.session.delete(job)
    db.session.commit()
    return jsonify({'success': True})


@jobs_bp.route('/export', methods=['GET'])
@login_required
def export_jobs():
    user_id = request.current_user.id
    fmt = request.args.get('format', 'csv')
    jobs = InterestedJob.query.filter_by(user_id=user_id).order_by(InterestedJob.collected_at.desc()).all()

    status_map = {'applied': '已投递', 'interviewing': '面试中', 'rejected': '已拒绝', 'interested': '有意向'}
    headers = ['公司', '岗位', '薪资', '地点', '经验要求', '学历要求', '状态', '匹配度', '投递时间']

    if fmt == 'json':
        data = [j.to_dict() for j in jobs]
        return Response(
            json.dumps(data, ensure_ascii=False, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': 'attachment; filename=boss-jobs-export.json'},
        )

    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = '投递记录'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4285F4', end_color='4285F4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row, j in enumerate(jobs, 2):
            values = [
                j.company_name, j.job_name, j.salary or '', j.location or '',
                j.experience or '', j.education or '',
                status_map.get(j.status, j.status), f"{j.match_score}分",
                j.created_at.strftime('%Y-%m-%d %H:%M') if j.created_at else '',
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=boss-jobs-export.xlsx'},
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for j in jobs:
        writer.writerow([
            j.company_name, j.job_name, j.salary or '', j.location or '',
            j.experience or '', j.education or '',
            status_map.get(j.status, j.status), f"{j.match_score}分",
            j.created_at.strftime('%Y-%m-%d %H:%M') if j.created_at else '',
        ])
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=boss-jobs-export.csv'},
    )
